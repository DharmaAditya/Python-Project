from typing import List, Dict
from database import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class Inventory:
    @staticmethod
    def add_item(name: str, quantity: int, price: float, category: str = None):
        """Menambahkan item baru"""
        if quantity < 0 or price < 0:
            raise ValueError("Quantity dan price tidak boleh negatif")
        add_item(name, quantity, price, category)
    
    @staticmethod
    def get_items() -> List[Dict]:
        """Mendapatkan semua item"""
        return get_all_items()
    
    @staticmethod
    def update_item(item_id: int, **kwargs):
        """Memperbarui item"""
        valid_fields = {'name', 'quantity', 'price', 'category'}
        for field in kwargs:
            if field not in valid_fields:
                raise ValueError(f"Field {field} tidak valid")
        
        if 'quantity' in kwargs and kwargs['quantity'] < 0:
            raise ValueError("Quantity tidak boleh negatif")
        
        if 'price' in kwargs and kwargs['price'] < 0:
            raise ValueError("Price tidak boleh negatif")
        
        update_item(item_id, **kwargs)
    
    @staticmethod
    def remove_item(item_id: int):
        """Menghapus item"""
        delete_item(item_id)
    
    @staticmethod
    def search(keyword: str) -> List[Dict]:
        """Mencari item"""
        return search_items(keyword)
    
    @staticmethod
    def get_low_stock(threshold: int = 5) -> List[Dict]:
        """Mendapatkan item dengan stok rendah"""
        items = get_all_items()
        return [item for item in items if item['quantity'] < threshold]
    
    @staticmethod
    def get_total_value() -> float:
        """Menghitung total nilai inventory"""
        items = get_all_items()
        return sum(item['quantity'] * item['price'] for item in items)

    @staticmethod
    def export_to_excel_pretty(filename: str = None):
        """Mengekspor data ke Excel dengan formatting"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"inventory_report_{timestamp}.xlsx"
        
        items = get_all_items()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Header
        headers = ["ID", "Nama Item", "Quantity", "Harga", "Total", "Kategori", "Terakhir Update"]
        ws.append(headers)
        
        # Style header
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        center_aligned = Alignment(horizontal='center')
        
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_aligned
        
        # Data rows
        for item in items:
            total_value = item['quantity'] * item['price']
            row = [
                item['id'],
                item['name'],
                item['quantity'],
                item['price'],
                total_value,
                item['category'] or "-",
                item['last_updated']
            ]
            ws.append(row)
        
        # Formatting data cells
        for row in ws.iter_rows(min_row=2, max_row=len(items)+1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.column in [3, 4, 5]:  # Quantity, Harga, Total
                    cell.alignment = Alignment(horizontal='right')
                    if cell.column in [4, 5]:  # Harga dan Total
                        cell.number_format = '#,##0.00'
        
        # Auto adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save file
        wb.save(filename)
        return filename